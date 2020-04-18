'''
Created on Sep 28, 2019

@author: Sasa33k

'''
import openpyxl
# Give the location of the file
my_path = "D:\Book1.xlsx"
my_wb_obj = openpyxl.load_workbook(my_path)
print(my_wb_obj.sheetnames)
ws = my_wb_obj["Sheet1"]
arr=[]
#my_wb_obj.active
for row in ws.rows:
    tmpArr = []
    for cell in row:
        print(cell.value)
        tmpArr.append(cell.value)
    arr.append(tmpArr)
"""      
my_row = my_sheet_obj.max_row
my_row = my_sheet_obj.max_row
for i in range(1, my_row + 1):
    cell_obj = my_sheet_obj.cell(row = i, column = 1)
    print(cell_obj.value)
"""

print(arr)
print(arr[0][1])



wb = openpyxl.Workbook()
ws = wb.active
for i in arr:
    ws.append(i)
wb.save('D:\create_sample.xlsx')