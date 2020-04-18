'''
Created on Sep 28, 2019

@author: Sasa33k

'''
import openpyxl

def load(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    print(wb.sheetnames)
    return wb

def dataSheet(wb, sheet):
    ws = wb[sheet]
    return ws

def dataRows(ws):
    arr=[]
    for row in ws.rows:
        tmpArr = []
        for cell in row:
            tmpArr.append(cell.value)
        if tmpArr[0]==None: break
        else: arr.append(tmpArr)
    return arr

def write(summary, data, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in summary:
        ws.append(i)
        
    for sData in data:
        print(sData[0])
        ws = wb.create_sheet(title = sData[0][0]) 
        for i in sData:
            ws.append(i)
        
    
    wb.save(path)


### Validate data template

def validateTitle(expTitle, dataTitle):
    #Validate Title: Sheet Names, Step Title
    
    if (all(elem in dataTitle  for elem in expTitle)):
        return(True)
    else:
        print("Template Title mismatch...")

def containElement(element, data):
    # Check if Row / column contain element
    for cell in data:
        if(cell.value == element):
            return True     
    return False


    
"""


#my_wb_obj.active
for row in ws.rows:
    tmpArr = []
    for cell in row:
        print(cell.value)
        tmpArr.append(cell.value)
    arr.append(tmpArr)
"""
"""      
my_row = my_sheet_obj.max_row
my_row = my_sheet_obj.max_row
for i in range(1, my_row + 1):
    cell_obj = my_sheet_obj.cell(row = i, column = 1)
    print(cell_obj.value)
"""

"""
print(arr)
print(arr[0][1])



wb = openpyxl.Workbook()
ws = wb.active
for i in arr:
    ws.append(i)
wb.save('D:\create_sample.xlsx')
"""